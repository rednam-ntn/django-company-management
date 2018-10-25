from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect, HttpResponse, Http404
from django.urls import reverse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
# permission_required,
from django.views.decorators.cache import never_cache
from django.db.models import Q

from .models import Project, Staff, Allocation, DateLog,\
    TBTC, TBTCManage, TBTCStatusLog, TBTCDocument, TBVT, TBVTManage
from dateutil.relativedelta import relativedelta
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.compat import range

today = dt.date.today()


def index(request):
    num_projects = Project.objects.all().count()
    num_staffs = Staff.objects.filter(d_out__isnull=True).count()
    return render(request, 'index.html',
                  {'num_projects': num_projects, 'num_staffs': num_staffs})

################################################


def d_range(start_date, end_date):
    for n in range(int((end_date - start_date).days)):
        yield start_date + dt.timedelta(days=n)


def time_round(time_in, inter):
    """
    time_in: datetime.timedelta()
    inter: interger type of round minutes
    """
    period = dt.timedelta(minutes=inter).total_seconds()
    remain = time_in.total_seconds() % period

    if remain >= period / 2:
        time_in = dt.timedelta(seconds=time_in.total_seconds()
                               + (period - remain))
        return time_in
    else:
        time_in = dt.timedelta(seconds=time_in.total_seconds() - remain)
        return time_in


def month_input_process(request):
    """
    PROCESS MONTH INPUT with:
        session.m_input, prev-input, next-input and month-input
    then return m_raw (unkown pattern) and m_input (a list as [year, month])
    """
    m_in_raw, m_input = None, None
    if "GET" == request.method:
        if request.session.get('m_input', False):
            m_in_raw = request.session['m_input']
            del request.session['m_input']
        else:
            m_input = [dt.datetime.now().year, dt.datetime.now().month]
    else:
        if request.POST.get('prev-input', False):
            m_in = request.POST['prev-input']
            m_conv = dt.datetime.strptime(str(m_in), '%m/%Y')
            m_input_dtb = m_conv - relativedelta(months=+1)
            m_input = [m_input_dtb.year, m_input_dtb.month]
        elif request.POST.get('next-input', False):
            m_in = request.POST['next-input']
            m_conv = dt.datetime.strptime(str(m_in), '%m/%Y')
            m_input_dtb = m_conv + relativedelta(months=+1)
            m_input = [m_input_dtb.year, m_input_dtb.month]
        elif request.POST.get('month-input', False):
            m_in_raw = request.POST['month-input']
        else:
            m_input = [dt.datetime.now().year, dt.datetime.now().month]
    return m_in_raw, m_input


def validate_m_in_raw(m_in_raw):
    """
    m_in_raw: month input by typing must follow pattern YYYY-mm
    """
    err_mess = None
    if (len(m_in_raw) < 6) or (len(m_in_raw) > 7):
        err_mess = "Nhập thông tin sai cấu trúc"
    if err_mess is None:
        try:
            m_input = m_in_raw.split('-')
            int(m_input[1])
            int(m_input[0])
        except ValueError:
            err_mess = "Thông tin nhập vào không phải là số"
        except IndexError:
            err_mess = "Thông tin nhập sai cấu trúc"
    if err_mess is None and int(m_input[1]) > 12:
        err_mess = "Chỉ có 12 tháng"
    return err_mess


def date_input_process(request):
    """
    PROCESS DATE INPUT with:
        session.d_input, prev-input, next-input and dateinput
    then and return date_db as datetime.date type that NOT GREATER THEN TODAY
    """
    if "GET" == request.method:
        if request.session.get('d_input', False):
            date_input = request.session['d_input']
            date_db = dt.datetime.strptime(date_input, '%Y-%m-%d').date()
            del request.session['d_input']
        else:
            date_db = today
    elif "POST" == request.method:
        if request.POST.get('prev-input', False):
            date_input = request.POST['prev-input']
            date_pydt = dt.datetime.strptime(date_input, '%Y-%m-%d').date()
            date_db = date_pydt - dt.timedelta(days=1)
        elif request.POST.get('next-input', False):
            date_input = request.POST['next-input']
            date_pydt = dt.datetime.strptime(date_input, '%Y-%m-%d').date()
            date_db = date_pydt + dt.timedelta(days=1)
        elif request.POST.get('dateinput', False):
            date_input = request.POST['dateinput']
            date_db = dt.datetime.strptime(date_input, '%Y-%m-%d').date()
        else:
            date_db = today
    if date_db > today:
        date_db = today
    return date_db


def m_report_cal(month, year, code):
    """
    month:      str(12)
    year:       str(2017)
    project_id: string type of project id
    """
    report = dict()
    if code == 'office':
        log_list = DateLog.objects.filter(
            project__isnull=True,
            date__month=month,
            date__year=year,
            )
    else:
        log_list = DateLog.objects.filter(
            project__id=code,
            date__month=month,
            date__year=year,
            )
    if log_list.exists():
        for log_d in log_list.order_by('staff__name').distinct('staff__name'):
            t_worked_day, t_sunday, t_over9 = 0, 0, 0
            t_bonus_h = dt.timedelta(seconds=0)
            for log in log_list:
                if log_d.staff == log.staff:

                    if log.shift1 or log.shift2:
                        if log.date.weekday() == 6:
                            t_sunday += 1  # Sunday
                        else:
                            t_worked_day += 1
                    if log.shift3 is not None:
                        t_bonus_h += log.shift3
                    if log.over9:
                        t_over9 += 1
                else:
                    continue
            # remove the zero status
                if (
                    t_worked_day == 0
                    and t_bonus_h == dt.timedelta(seconds=0)
                    and t_sunday == 0
                    and t_over9 == 0
                ) is not True:
                    report[log.staff] = [
                            t_worked_day, t_bonus_h, t_sunday, t_over9]
    return report
########################


shifts = {
            'shift1': 'ca sáng',
            'shift2': 'ca chiều',
            'shift3': 'tăng ca',
            'over9': 'trực đêm',
            }


def create_staff_datelog(staff_list, date, site=None):
    """
    staff_list: an QuerySet/list of staff that need to search DateLog in DB
    date:       date to search in DB
    site:       project object, ignore if site is Office
    """
    staffs = dict()
    for staff in staff_list:
        log = {'date': date, 'staff': staff, 'project': site}
        try:
            log_list = DateLog.objects.get(**log)
        except DateLog.DoesNotExist:
            log_list = log

        staffs[staff] = log_list
    return staffs


def update_and_save_checked_log(staffs, checkeds, date, site=None):
    """
    staffs:     a dict that return from create_staff_datelog() above
    checkeds:   a request.POST.dict()
    """
    # create a stock/base datelog for staffs
    staffs_input_log = dict()
    for staff in staffs:
        if isinstance(staffs[staff], dict) is not True:
            if site is None:
                staffs_input_log[staff] = {
                    'date': date, 'shift1': False, 'shift2': False}
            else:
                staffs_input_log[staff] = {
                    'date': date, 'over9': False,
                    'shift1': False, 'shift2': False}

    # update input_checked to staffs_input_log
    for checked, status in checkeds.items():
        if '-_-=-_-' in checked:
            log = {}
            shift = checked.split('-_-=-_-')[0]
            if shift != "shift3" and status == "on":
                log[shift] = True
            elif shift == "shift3":
                if ':' in status:
                    hr = int(status.split(':')[0])
                    min = int(status.split(':')[1])
                    log['shift3'] = dt.timedelta(hours=hr, minutes=min)
                elif status == '' or status == ' ':
                    log['shift3'] = None
            staff = Staff.objects.get(id=checked.split('-_-=-_-')[1])
            if staff not in staffs_input_log:
                staffs_input_log[staff] = log
            else:
                staffs_input_log[staff].update(log)

    # SAVE to Database
    for staff, input_log in staffs_input_log.items():
        if isinstance(staffs[staff], dict):
            staffs[staff] = DateLog.objects.create(
                date=date, staff=staff, project=site)
        staffs[staff].__dict__.update(input_log)
        staffs[staff].save()

################################################


@login_required
def site_list(request):
    if "GET" == request.method:
        project_list = Project.objects.all()
        projects = dict()
        for project in project_list:
            p_allos = Allocation.objects.filter(project=project.id,
                                                d_out__isnull=True)
            pos_count = [0, 0]
            for p_allo in p_allos:
                if str(p_allo.staff.position.short_text) == 'GS':
                    pos_count[0] += 1
                elif str(p_allo.staff.position.short_text) == 'CN' or\
                        str(p_allo.staff.position.short_text) == 'ĐT':
                    pos_count[1] += 1

            projects[project] = pos_count
            o_staff_count = Staff.objects.filter(
                office=True, d_out__isnull=True).count()

        return render(request, 'summary/site_list.html', {
            "projects": projects, "o_staff_count": o_staff_count})
    else:
        try:
            site = request.POST['site']
        except (KeyError):
            return render(request, 'summary/site_list.html',
                          {'err_mess': "Thiếu thông tin Nơi làm việc"})
        if request.POST.get('m_input', False):
            m_input = request.POST['m_input']
            request.session['m_input'] = m_input
            if site == "office":
                return redirect('o_m_report')
            else:
                return redirect('p_m_report', project_id=site)
        elif request.POST.get('d_input', False):
            d_input = request.POST['d_input']
            request.session['d_input'] = d_input
            if site == "office":
                return redirect('o_d_report')
            else:
                return redirect('p_d_report', project_id=site)
        else:
            return render(request, 'summary/site_list.html',
                          {'err_mess': "Thiếu thông tin thời gian"})


@login_required
def staff_list(request):
    # PROJECT SITE
    staff_list = Staff.objects.filter(office=False, d_out__isnull=True)
    p_staffs = dict()
    for staff in staff_list:
        # get project_checked
        s_allos = Allocation.objects.filter(staff=staff.id, d_out__isnull=True)
        if s_allos.exists():
            projects_checked = dict()
            for s_allo in s_allos:
                log_list = DateLog.objects.filter(
                    staff=s_allo.staff,
                    project=s_allo.project,
                    date=today
                )
                if log_list.exists():
                    projects_checked[s_allo.project] = log_list
                else:
                    projects_checked[s_allo.project] = ''
            p_staffs[staff] = projects_checked

    # OFFICE
    staff_list = Staff.objects.filter(office=True, d_out__isnull=True)
    o_staffs = dict()
    for staff in staff_list:
        # get project_checked
        log_list = DateLog.objects.filter(
            staff=staff,
            date=today
        )
        if log_list.exists():
            o_staffs[staff] = log_list
        else:
            o_staffs[staff] = ''

    return render(request, 'summary/staff/list.html', {
        "p_staffs": p_staffs,
        "o_staffs": o_staffs})


@user_passes_test(lambda u: u.is_staff)
def staffs_history(request):
    date_db = date_input_process(request)

    # PROJECT SITE
    staff_list = Staff.objects.filter(
        Q(office=False),
        Q(d_in__lte=date_db),
        (Q(d_out__isnull=True) | Q(d_out__gte=date_db)),
        )
    p_staffs = dict()
    for staff in staff_list:
        p_staffs[staff] = Allocation.objects.filter(
            Q(staff=staff.id),
            Q(d_in__lte=date_db),
            Q(d_out__isnull=True) | Q(d_out__gte=date_db),
            )

    # OFFICE
    o_staffs = Staff.objects.filter(
        Q(office=True),
        Q(d_in__lte=date_db),
        (Q(d_out__isnull=True) | Q(d_out__gte=date_db)),
        )

    pos_count = dict()
    for staff in Staff.objects.filter(
        Q(d_in__lte=date_db),
        (Q(d_out__isnull=True) | Q(d_out__gte=date_db)),
            ):
            pos_count[staff.position] = pos_count.get(staff.position, 0) + 1

    return render(request, 'summary/staff/history.html', {
        "p_staffs": p_staffs, "o_staffs": o_staffs,
        "date_db": date_db, "pos_count": pos_count})


@login_required
def office_detail(request):
    staff_list = Staff.objects.filter(office=True, d_out__isnull=True)
    staffs = dict()
    for staff in staff_list:
        # get project_checked
        log_list = DateLog.objects.filter(
            staff=staff,
            date=today
        )
        if log_list.exists():
            staffs[staff] = log_list
        else:
            staffs[staff] = ''

    return render(request, 'summary/office/detail.html', {
        "staffs": staffs})


@login_required
def project_detail(request, project_id):
    try:
        project = get_object_or_404(Project, pk=project_id)
    except Project.DoesNotExist:
        raise Http404("Project does not exists")
    # Position today count
    pos_count = [0, 0]
    for p_allo in Allocation.objects.filter(
            project=project.id, d_out__isnull=True):
        if str(p_allo.staff.position.short_text) == 'GS':
            pos_count[0] += 1
        elif str(p_allo.staff.position.short_text) == 'CN' or\
                str(p_allo.staff.position.short_text) == 'ĐT':
            pos_count[1] += 1

    # filter all Allocation by project
    staffs_info = dict()
    for p_allo in Allocation.objects.filter(project=project.id):
        if p_allo.d_out:
            staffs_info[p_allo.staff] = [p_allo.d_in, p_allo.d_out]
        else:
            staffs_info[p_allo.staff] = [p_allo.d_in]
    # TBTC list now and past
    p_tbtc_now = TBTCManage.objects.filter(
        Q(d_receive__project=project),
        Q(d_receive__date_crt__lte=today),
        (Q(d_return__isnull=True) | Q(d_return__date_crt__gte=today)),
        Q(unit__d_in__lte=today),
        (Q(unit__d_out__isnull=True) | Q(unit__d_out__gte=today)),
        )
    p_tbtc = TBTCManage.objects.filter(
        Q(d_receive__project=project),
        Q(d_receive__date_crt__lte=today),
        Q(d_return__date_crt__lt=today),
        Q(unit__d_in__lte=today), Q(unit__d_out__lt=today)
        )

    # TBVT list
    p_tbvt = TBVTManage.objects.filter(document__project=project)

    return render(request, 'summary/project/detail.html', {
        "project": project,
        "staffs_info": staffs_info,
        "pos_count": pos_count,
        "p_tbtc": p_tbtc,
        "p_tbtc_now": p_tbtc_now,
        "p_tbvt": p_tbvt,
        })


@login_required
def staff_detail(request, staff_id):
    try:
        staff = get_object_or_404(Staff, pk=staff_id)
    except Project.DoesNotExist:
        raise Http404("Project does not exists")

    project_infos = dict()
    # filt all Allocation by staff
    s_allos = Allocation.objects.filter(staff=staff.id)
    for s_allo in s_allos:
        if s_allo.d_out:
            project_infos[s_allo.project] = (s_allo.d_in, s_allo.d_out)
        else:
            project_infos[s_allo.project] = (s_allo.d_in)

    return render(request, 'summary/staff/detail.html', {
        "staff": staff,
        "project_infos": project_infos,
        })

########################


@never_cache
@user_passes_test(lambda u: u.has_perm('summary.office_check'))
def office_check(request):
    date_db = date_input_process(request)
    staff_list = Staff.objects.filter(
        Q(office=True), Q(d_in__lte=date_db),
        (Q(d_out__isnull=True) | Q(d_out__gt=date_db)))

    if staff_list.exists() is not True:
        return render(request, 'summary/office/check.html', {
            'date_db': date_db,
            'err_mess': "Không có nhân sự nào làm việc ở Văn phòng"})

    staffs = create_staff_datelog(staff_list, date=date_db, site=None)
    if "GET" == request.method:
        return render(request, 'summary/office/check.html', {
            'date_db': date_db, "staffs": staffs})
    else:
        checkeds = request.POST.dict()
        # validate input_checked
        input_logs = 0
        for checked, status in checkeds.items():
            if "-_-=-_-" in checked:
                input_logs += 1
                if checked.split('-_-=-_-')[0] == "shift3" and ':' in status:
                    try:
                        int(status.split(':')[0])
                        int(status.split(':')[1])
                    except ValueError:
                        return render(
                            request, 'summary/office/check.html',
                            {'err_mess': "Vui lòng nhập thời gian ở dạng số"})

        if input_logs > 0:
            update_and_save_checked_log(staffs, checkeds, date_db, site=None)
        return render(request, 'summary/office/check.html', {
            'date_db': date_db, "staffs": staffs})


@never_cache
@login_required
def project_check(request, project_id):
    try:
        project = get_object_or_404(Project, pk=project_id)
    except Project.DoesNotExist:
        return HttpResponseRedirect(reverse("site_list"))

    cur_user = User.objects.get(username=request.user.username)
    if cur_user.is_superuser is not True:
        if cur_user not in project.allowed_users.all():
            return render(request, 'summary/project/check.html', {
                'err_mess': "Bạn không có quyền truy cập chấm công cho " +
                "dự án này, vui lòng 'Thoát' ra và 'Đăng nhập' " +
                "bằng tài khoản hợp lệ."})

    date_db = date_input_process(request)
    # Search for if any staff in project
    p_allos = Allocation.objects.filter(
        Q(project=project.id), Q(d_in__lte=date_db),
        (Q(d_out__isnull=True) | Q(d_out__gt=date_db)))
    if p_allos.exists() is not True:
        return render(request, 'summary/project/check.html', {
            'err_mess': "Chưa có nhân sự nào được phân công"})

    staff_list = [p_allo.staff for p_allo in p_allos]
    staffs = create_staff_datelog(staff_list, date=date_db, site=project)

    if "GET" == request.method:
        return render(request, 'summary/project/check.html', {
            "project": project, "staffs": staffs, 'date_db': date_db})

    else:
        checkeds = request.POST.dict()
        # validate input_checked
        input_logs = 0
        for checked, status in checkeds.items():
            if "-_-=-_-" in checked:
                input_logs += 1
                if checked.split('-_-=-_-')[0] == "shift3" and ':' in status:
                    try:
                        int(status.split(':')[0])
                        int(status.split(':')[1])
                    except ValueError:
                        return render(
                            request, 'summary/project/check.html',
                            {'err_mess': "Vui lòng nhập thời gian ở dạng số"})
        if input_logs > 0:
            update_and_save_checked_log(
                staffs, checkeds, date=date_db, site=project)
        return render(request, 'summary/project/check.html', {
            "project": project, "staffs": staffs, 'date_db': date_db})

        return HttpResponseRedirect(reverse("project_check",
                                            args=[project_id]))


########################


@login_required
def o_m_report(request):
    # validate month_input_raw text
    m_in_raw, m_input = month_input_process(request)
    if m_in_raw is not None:
        err_mess = validate_m_in_raw(m_in_raw)
        if err_mess is not None:
            return render(request, 'summary/office/monthreport.html', {
                    'err_mess': err_mess})
        else:
            m_input = m_in_raw.split('-')

    month_display = str(m_input[1]) + '/' + str(m_input[0])
    staffs_timeinfo = m_report_cal(m_input[1], m_input[0], code='office')
    return render(request, 'summary/office/monthreport.html', {
            "month_display": month_display,
            "staffs_timeinfo": staffs_timeinfo})


@login_required
def p_m_report(request, project_id):
    try:
        project = get_object_or_404(Project, pk=project_id)
    except Project.DoesNotExist:
        return HttpResponseRedirect(reverse("site_list"))

    # validate month_input_raw text
    m_in_raw, m_input = month_input_process(request)
    if m_in_raw is not None:
        err_mess = validate_m_in_raw(m_in_raw)
        if err_mess is not None:
            return render(request, 'summary/office/monthreport.html', {
                    'err_mess': err_mess})
        else:
            m_input = m_in_raw.split('-')

    month_display = str(m_input[1]) + '/' + str(m_input[0])
    staffs_timeinfo = m_report_cal(m_input[1], m_input[0], code=project_id)
    return render(
        request,
        'summary/project/monthreport.html', {
            "project": project, "month_display": month_display,
            "staffs_timeinfo": staffs_timeinfo})


################################################


def style_range(ws, cell_range, border=Border(), fill=None, font=None,):
    """
    Apply styles to a range of cells as if they were a single cell.
    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        le = row[0]
        ri = row[-1]
        le.border = le.border + left
        ri.border = ri.border + right
        if fill:
            for c in row:
                c.fill = fill


def export_mreport_excel(request):
    if "GET" == request.method:
        return HttpResponseRedirect(reverse("site_list"))
    else:
        if request.POST.get('month-export', False) and \
                request.POST.get('project_id', False):
            try:
                m_in_raw = request.POST['month-export']
                m_input = m_in_raw.split('/')
                int(m_input[0])
                int(m_input[1])
            except ValueError:
                return HttpResponseRedirect(reverse("site_list"))
        else:
            return HttpResponseRedirect(reverse("site_list"))

    month_in = m_input[0]
    year_in = m_input[1]
    project_id = request.POST['project_id']
    month_display = str(month_in) + '/' + str(year_in)
    if project_id != 'office':
        project = Project.objects.get(id=project_id)
        name = str(project.name) + ', ' + str(project.location)
    else:
        name = 'Văn Phòng'

    staffs_timeinfo = m_report_cal(month_in, year_in, project_id)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="mreport.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = str(month_in) + '-' + str(year_in)

    bold = Font(bold=True)
    ws.append(['TỔNG KẾT THÁNG:', month_display])
    ws.append(['ĐỊA ĐIỂM:', name])
    ws['B1'].font = bold
    ws['B2'].font = bold
    ws.append([''])
    ws.append([
        'Nhân viên',
        'Vị trí',
        'Bộ Phận',
        'Ngày công',
        'Tổng giờ làm tăng ca',
        'Tổng ngày CN',
        'Tổng đêm trực'])
    for staff, times in staffs_timeinfo.items():
        row_data = list()
        row_data.append(str(staff.name))
        row_data.append(str(staff.position))
        row_data.append(str(staff.deparment))
        row_data.append(times[0])   # t_worked_day
        row_data.append(times[1])   # t_bonus_h
        row_data.append(times[2])   # t_sunday
        row_data.append(times[3])   # t_over9
        ws.append(row_data)

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="DDDDDD")

    style_range(ws, 'A4:G' + str(4 + len(staffs_timeinfo)), border=border)
    style_range(ws, 'A4:G4', border=border, fill=fill)

    wb.save(response)
    return response

################################################


def tbtc_date_list(date):
    """
    info:      an empty dict() that receive:
        key =       tb
        value =     queryset of that tb in TBTCManage
    site_count:      an empty dict() that receive:
        key =       site, include "Kho"
        value =     tb.count from TBTCManage
    type_count:      an empty dict() that receive:
        key =       type of TBTC
        value =     tb.count of that TBTC from TBTCManage
    date:   date format (yyyy/mm/dd)
    """
    info = dict()
    site_count = dict()
    site_list = dict()
    type_count = dict()
    status = dict()
    for project in Project.objects.all():
        site_count[project] = TBTCManage.objects.filter(
            Q(d_receive__project=project),
            Q(d_receive__date_crt__lte=date),
            (Q(d_return__isnull=True) | Q(d_return__date_crt__gt=date)),
            Q(unit__d_in__lte=date),
            (Q(unit__d_out__isnull=True) | Q(unit__d_out__gt=date)),
            ).count()
        # site_list for status multiple select
        site_list[project] = TBTCManage.objects.filter(
            Q(d_receive__project=project),
            Q(d_receive__date_crt__lte=date),
            (Q(d_return__isnull=True) | Q(d_return__date_crt__gt=date)),
            Q(unit__d_in__lte=date),
            (Q(unit__d_out__isnull=True) | Q(unit__d_out__gt=date)),
            )
    site_count['Kho'] = 0

    tbtc_list = TBTC.objects.filter(
        Q(d_in__lte=date),
        (Q(d_out__isnull=True) | Q(d_out__gte=date)),
        )
    status['Không sử dụng được'] = 0
    status['Còn sử dụng được'] = 0

    kho = list()
    for tb in tbtc_list:
        # Get Repair list:
        if str(tb.status())[0] == 'H':
            status['Không sử dụng được'] += 1
        else:
            status['Còn sử dụng được'] += 1

        # Get TBTC Manage
        type_count[(tb.type, tb.brand)] =\
            type_count.get((tb.type, tb.brand), 0) + 1

        info[tb] = TBTCManage.objects.filter(
            Q(unit=tb),
            Q(d_receive__date_crt__lte=date),
            (Q(d_return__isnull=True) |
             Q(d_return__date_crt__gte=date)),
            )
        if len(info[tb]) < 1 or \
                len(info[tb].filter(
                    d_return__MaVB__contains='NK',
                    d_return__date_crt=date
                    )) > 0:
            site_count['Kho'] += 1
            kho.append(tb)

    site_list['Kho'] = kho
    return info, site_count, site_list, type_count, status


def number_of_next_tb_doc(type):
    """
    type:       string that define VBNK or VBBG
    vb_num:     number of the next vbtbtc
    """
    vb_count = TBTCDocument.objects.filter(
        MaVB__contains=type
        ).count() + 1
    if len(str(vb_count)) == 1:
        vb_num = "00" + str(vb_count)
    elif len(str(vb_count)) == 2:
        vb_num = "0" + str(vb_count)
    else:
        vb_num = str(vb_count)
    return vb_num


def validate_tb_doc(tb, MaVB):
    """
    tb:     TBTC object
    MaVB:   TBTCDocument.MaVB and is string type
    """
    if "VBBG" in MaVB:
        try:
            cur_sent_doc = TBTCManage.objects.filter(
                unit=tb, d_return__isnull=True
                ).latest("id").d_receive.MaVB
        except TBTCManage.DoesNotExist:
            return ["G"]
        else:
            return ["B", "BG", tb.MaTBTC, cur_sent_doc]
    elif "VBNK" in MaVB or "VBCT" in MaVB:
        try:
            cur_sent_doc = TBTCManage.objects.filter(
                unit=tb, d_return__isnull=True
                ).latest("id").d_receive.MaVB
        except TBTCManage.DoesNotExist:
            return ["B", "NK/CT", tb.MaTBTC, ""]
        else:
            return ["G", "", ""]


def update_or_save_doc_to_tbmanage(tb, MaVB):
    """
    tb:     TBTC object
    MaVB:   TBTCDocument.MaVB and is string type
    """
    if "VBBG" in MaVB:
        manage = {
            'unit': tb,
            'd_receive': TBTCDocument.objects.get(MaVB=MaVB),
        }
        manage_event = TBTCManage(**manage)
        manage_event.save()
    elif "VBNK" in MaVB:
        cur_doc = TBTCManage.objects.filter(
            unit=tb, d_return__isnull=True
            ).latest("id").d_receive.MaVB
        m_update = {
            'd_return': TBTCDocument.objects.get(MaVB=MaVB)}
        manage = {
            'unit': tb,
            'd_receive': TBTCDocument.objects.get(MaVB=cur_doc),
        }
        TBTCManage.objects.filter(**manage).update(**m_update)
    elif "VBCT" in MaVB:
        cur_doc = TBTCManage.objects.filter(
            unit=tb, d_return__isnull=True
            ).latest("id").d_receive.MaVB
        m_update = {
            'd_return': TBTCDocument.objects.get(MaVB=MaVB)}
        old_manage = {
            'unit': tb,
            'd_receive': TBTCDocument.objects.get(
                MaVB=cur_doc),
        }
        TBTCManage.objects.filter(**old_manage).update(**m_update)
        new_manage = {
            'unit': tb,
            'd_receive': TBTCDocument.objects.get(MaVB=MaVB),
        }
        manage_event = TBTCManage(**new_manage)
        manage_event.save()


def tbtc_status(unit, date=None):
    if date is None:
        tb_stt = TBTCStatusLog.objects.filter(unit=unit)
        if tb_stt.exists():
            tb_stt = tb_stt.first()
            if tb_stt.status is True:
                date_display = dt.datetime.strftime(tb_stt.date, '%d/%m/%Y')
                stt = 'Hư hỏng từ ngày ' + date_display
            else:
                stt = 'Tốt'
        else:
            stt = 'Tốt'
        return stt
    else:
        tb_stt = TBTCStatusLog.objects.filter(unit=unit, date__lte=date)
        if tb_stt.exists():
            tb_stt = tb_stt.first()
            if tb_stt.status is True:
                date_display = dt.datetime.strftime(tb_stt.date, '%d/%m/%Y')
                stt = 'Hư hỏng từ ngày ' + date_display
            else:
                stt = 'Tốt'
        else:
            stt = 'Tốt'
        return stt

#####################


@user_passes_test(lambda u: u.has_perm('summary.manage_TBTC'))
def tbtc_list(request):
    date_db = date_input_process(request)
    if 'POST' == request.method:
        if request.POST.get('next-doc', False):
            completed_list = list()
            err_list = list()
            for tb_doc in request.POST.getlist('next-doc', []):
                if tb_doc != "None":
                    MaTB, MaVB = tb_doc.split('=')[0], tb_doc.split('=')[1]
                    validation = validate_tb_doc(
                        TBTC.objects.get(MaTBTC=MaTB), MaVB)
                    if validation[0] == "G":
                        update_or_save_doc_to_tbmanage(
                            TBTC.objects.get(MaTBTC=MaTB), MaVB)
                        completed_list.append([MaTB, MaVB])
                    else:
                        err_list.append(validation[1:4])
            if len(err_list) == 0 and len(completed_list) == 0:
                err_list = ("Chưa có Thiết bị nào được cập nhật Văn bản",)
            return render(request, 'summary/tbtc/list.html', {
                "completed_list": completed_list,
                "err_list": err_list,
                })
        else:
            repair_list = request.POST.getlist('tbtc-repair', [])
            msg_list = list()
            for tb_id in repair_list:
                tb = TBTC.objects.get(id=tb_id)
                if str(tb.status())[0] == 'T':
                    log = dict()
                    log['unit'] = tb
                    log['status'] = True
                    log['date'] = today
                    event = TBTCStatusLog(**log)
                    event.save()
                    msg_list.append(tb)

            for tb in TBTC.objects.filter(
                    Q(d_in__lte=today),
                    (Q(d_out__isnull=True) | Q(d_out__gte=today)),
                    ):
                if str(tb.status())[0] == 'H'\
                 and str(tb.id) not in repair_list:
                    log = dict()
                    log['unit'] = tb
                    log['status'] = False
                    log['date'] = today
                    event = TBTCStatusLog(**log)
                    event.save()
                    msg_list.append(tb)
            if len(msg_list) > 0:
                return render(request, 'summary/tbtc/list.html', {
                    "msg_list": msg_list
                    })

    doc_list = TBTCDocument.objects.filter(
        date_crt__gte=(date_db - relativedelta(days=7)))
    tbtc_info, tbtc_site_count, tbtc_site_list, tbtc_type_count, tbtc_status =\
        tbtc_date_list(date_db)

    return render(request, 'summary/tbtc/list.html', {
        "date_db": date_db, "doc_list": doc_list,
        "tbtc_info": tbtc_info, "tbtc_site_count": tbtc_site_count,
        "tbtc_site_list": tbtc_site_list, "tbtc_type_count": tbtc_type_count,
        "tbtc_status": tbtc_status})


@user_passes_test(lambda u: u.has_perm('summary.manage_TBTC'))
def tbtc_manage(request, TBTC_id):
    try:
        tb = get_object_or_404(TBTC, pk=TBTC_id)
    except TBTC.DoesNotExist:
        return HttpResponseRedirect(reverse("tbtc_list"))

    manages_raw = TBTCManage.objects.filter(unit=tb)
    project_list = Project.objects.all()
    sponsor_list = Staff.objects.filter(
                (Q(deparment__short_text='BT') | Q(deparment__short_text='TT'))
                & Q(d_out__isnull=True) & ~Q(position__short_text='CN'))
    manage_list = list()
    for manage in manages_raw:
        manage_list.append([
            manage.d_receive.MaVB, manage.d_receive.project,
            manage.d_receive.sponsor, manage.d_receive.date_crt])
        if manage.d_return:
            if manage.d_return.project == manage.d_receive.project \
                    or manage.d_return.project is None:
                manage_list.append([
                    manage.d_return.MaVB, "Kho",
                    manage.d_return.sponsor, manage.d_return.date_crt
                    ])
    if len(manage_list) == 0:
        cur_manage = ["--", "Kho", "--"]
    else:
        cur_manage = manage_list[-1]

    render_index = {
        "manage_list": manage_list, "project_list": project_list,
        "sponsor_list": sponsor_list, "cur_manage": cur_manage, "tb": tb}
    if "POST" == request.method:
        if request.POST.get('next-site', False)\
         and request.POST.get('next-sponsor', False):
            site = request.POST['next-site']
            sponsor = request.POST['next-sponsor']
            if site == cur_manage[1]:
                err_mess = "Địa điểm chuyển đi trùng với Địa điểm của Thiết bị"
                render_index["err_mess"] = err_mess
            else:
                document = {
                    'sponsor': Staff.objects.get(name=sponsor),
                    'date_crt': today}
                if site == "Kho":  # Nhập Kho từ Công trình
                    MaVB = "VBNK - " + number_of_next_tb_doc("NK")
                    document.update({
                        'MaVB': MaVB,
                        'name': "Văn bản Nhập kho",
                        'project': Project.objects.get(name=cur_manage[1])})
                    TBTCDocument(**document).save()
                    update_or_save_doc_to_tbmanage(tb, MaVB)
                elif cur_manage[1] == "Kho":  # Bàn giao ra Công trình từ Kho
                    MaVB = "VBBG - " + number_of_next_tb_doc("BG")
                    document.update({
                        'MaVB': MaVB,
                        'name': "Cấp Phát TBTC",
                        'project': Project.objects.get(name=site)})
                    TBTCDocument(**document).save()
                    update_or_save_doc_to_tbmanage(tb, MaVB)
                else:  # Chuyển tiếp TB từ Công trình này sang Công trình khác
                    MaVB = "VBCT - " + number_of_next_tb_doc("CT")
                    document.update({
                        'MaVB': MaVB,
                        'name': "Chuyển tiếp Thiết bị",
                        'project': Project.objects.get(name=site)})
                    TBTCDocument(**document).save()
                    update_or_save_doc_to_tbmanage(tb, MaVB)
            return HttpResponseRedirect(reverse('tbtc_manage', args=[TBTC_id]))

        else:
            return HttpResponseRedirect(reverse("tbtc_list"))

    return render(request, 'summary/tbtc/manage.html', render_index)


def export_tbtc_excel(request):  # NEED TO CHECK AGAIN
    if "GET" == request.method:
        return HttpResponseRedirect(reverse("tbtc_list"))
    else:
        if request.POST.get('date-export', False):
            try:
                date_db = request.POST['date-export']
            except ValueError:
                return HttpResponseRedirect(reverse("tbtc_list"))
        else:
            return HttpResponseRedirect(reverse("tbtc_list"))

    date_display = dt.datetime.strftime(date_db, '%d/%m/%Y')
    tbtc_info, tbtc_site_count, tbtc_site_list, tbtc_type_count, tbtc_status =\
        tbtc_date_list(date_db)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] =\
        'attachment; filename="ThongKeTBTC.xlsx"'

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = 'DanhSach'

    bold = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="DDDDDD")
    sumfill = PatternFill("solid", fgColor="cccccc")

    ws_info.append(['DANH SÁCH THIẾT BỊ THI CÔNG'])
    ws_info.append(['Ngày:', date_display])
    ws_info['A1'].font = bold
    ws_info['B2'].font = bold
    ws_info.append([''])
    ws_info.append([
        'Mã Thiết bị',
        'Tên',
        'Ký hiệu',
        'Nhãn hiệu',
        'Thời gian sử dụng',
        'Công trình',
        'Người quản lý',
        'Văn bản giao',
        'Văn bản trả về kho'])
    for tb, infos in tbtc_info.items():
        row_data = list()
        row_data.append(str(tb.MaTBTC))
        row_data.append(str(tb.type))
        row_data.append(str(tb.type.short_text))
        row_data.append(str(tb.brand))
        row_data.append(tb.day())
        if len(infos) > 0:
            for info in infos:
                row_data.append(str(info.d_receive.project))
                row_data.append(str(info.d_receive.sponsor))
                row_data.append(
                    str(info.d_receive) +
                    ', ngày ' +
                    str(dt.datetime.strftime(
                        info.d_receive.date_crt,
                        '%d/%m/%Y'
                        ))
                    )
                if info.d_return is None:
                    row_data.append(' ')
                else:
                    row_data.append(
                        str(info.d_return) +
                        ', ngày ' +
                        str(dt.datetime.strftime(
                            info.d_return.date_crt,
                            '%d/%m/%Y'
                            ))
                        )
        else:
            row_data.append('Kho')
            row_data.append(' ')
            row_data.append(' ')
            row_data.append(' ')

        ws_info.append(row_data)
    style_range(ws_info, 'A4:I' + str(4 + len(tbtc_info)), border=border)
    style_range(ws_info, 'A4:I4', border=border, fill=fill)

    ws_s_count = wb.create_sheet("TKCongTrinh")
    ws_s_count.append(['THỐNG KÊ THEO CÔNG TRÌNH'])
    ws_s_count.append(['Ngày:', date_display])
    ws_s_count['A1'].font = bold
    ws_s_count['B2'].font = bold
    ws_s_count.append([''])
    ws_s_count.append(['Công trình', 'Số lượng'])
    len_c = 0
    for p, count in tbtc_site_count.items():
        if count > 0:
            ws_s_count.append([str(p), int(count)])
            len_c += 1
    ws_s_count.append(['Tổng', '=SUM(B5:B' + str(4 + len_c) + ')'])
    style_range(ws_s_count, 'A4:B' + str(4 + len_c),
                border=border)
    style_range(ws_s_count, 'A4:B4', border=border, fill=fill)
    style_range(ws_s_count, 'A' + str(5 + len_c) + ':B' + str(5 + len_c),
                border=border, fill=sumfill)

    ws_t_count = wb.create_sheet("TKThietBi")
    ws_t_count.append(['THỐNG KÊ THEO THIẾT BỊ'])
    ws_t_count.append(['Ngày:', date_display])
    ws_t_count['A1'].font = bold
    ws_t_count['B2'].font = bold
    ws_t_count.append([''])
    ws_t_count.append(['Mã TB', 'Tên TB', 'Số lượng'])
    len_c = 0
    for tb, count in tbtc_type_count.items():
        if count > 0:
            ws_t_count.append([
                str(tb[0].short_text) + '.' + str(tb[1].short_text),
                str(tb[0]) + '.' + str(tb[1]),
                int(count)
                 ])
            len_c += 1
    ws_t_count.append(['Tổng', '', '=SUM(C5:C' + str(4 + len_c) + ')'])
    style_range(ws_t_count, 'A4:C' + str(4 + len_c),
                border=border)
    style_range(ws_t_count, 'A4:C4', border=border, fill=fill)
    style_range(ws_t_count, 'A' + str(5 + len_c) + ':C' + str(5 + len_c),
                border=border, fill=sumfill)

    wb.save(response)
    return response

################################################


@user_passes_test(lambda u: u.has_perm('summary.manage_TBVT'))
def tbvt_list(request):
    date_db = date_input_process(request)
    if date_db > today:
        date_db = today

    tbvt_site_list = dict()
    tbvt_list = TBVT.objects.all()
    for project in Project.objects.all():
        tbvt_sum = dict()
        tbvt_list_sorted = list()
        for vt_manage in TBVTManage.objects.filter(
            Q(document__status='ex'),
            Q(document__project=project),
            Q(document__date_crt__lte=date_db),
                ):
            tbvt_sum[vt_manage.unit.MaTBVT] =\
             tbvt_sum.get(vt_manage.unit.MaTBVT, 0) + vt_manage.quantity
        for vt in tbvt_list:
            if vt.MaTBVT not in tbvt_sum:
                tbvt_sum[vt.MaTBVT] = 0
        for key in sorted(tbvt_sum.keys()):
            tbvt_list_sorted.append(tbvt_sum[key])
        tbvt_site_list[project] = tbvt_list_sorted

    code_vt_list = list()
    tbvt_sorted_list = list()
    for vt in tbvt_list:
        code_vt_list.append(str(vt.MaTBVT))
    code_vt_list.sort()
    for code in code_vt_list:
        tbvt_sorted_list.append([code, TBVT.objects.get(MaTBVT=code).quantity])

    return render(request, 'summary/tbvt/list.html', {
        "date_db": date_db,
        "tbvt_site_list": tbvt_site_list, "tbvt_sorted_list": tbvt_sorted_list,
        })

################################################


@user_passes_test(lambda u: u.is_superuser)
def datelog_full(request):
    m_in_raw, m_input = month_input_process(request)
    if m_in_raw is not None:
        err_mess = validate_m_in_raw(m_in_raw)
        if err_mess is not None:
            return render(request, 'summary/office/monthreport.html', {
                    'err_mess': err_mess})
        else:
            m_input = m_in_raw.split('-')

    month_display = str(m_input[1]) + '/' + str(m_input[0])
    staffs_timeinfo = m_report_cal(m_input[1], m_input[0], code='office')
    return render(request, 'summary/office/monthreport.html', {
            "month_display": month_display,
            "staffs_timeinfo": staffs_timeinfo})
    """
    if "GET" == request.method:
        log_list = DateLog.objects.all()
        return render(request, 'summary/datelog_full.html', {
            "log_list": log_list,
            })
    else:
        checked_time = request.POST.lists()
        event = request.POST.dict()
        return render(request, 'summary/datelog_full.html', {
            "checked_time": checked_time,
            "event": event,
            })
    """
